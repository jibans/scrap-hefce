from selenium import webdriver
import unicodedata
import time
from openpyxl import load_workbook
from openpyxl.writer.write_only import WriteOnlyCell

def getDataPerpage(page_link):
    print("Sub page")
    wb = load_workbook(filename='All UK HE Providers.xlsx')
    ws = wb['ALL HE PROVIDERS']
    row_no=1
    driver = webdriver.Chrome("C:/chromedriver_win32/chromedriver.exe")
    for link in page_link:
        driver.get(link)
        for li in driver.find_elements_by_xpath('//ul[@id="browseProviderResults"]//li'):
            anchors = li.find_elements_by_tag_name('a')
            for a in anchors:
                link= driver.find_element_by_link_text(a.text)
                web_link=unicodedata.normalize('NFKD', link.get_attribute('href')).encode('ascii','ignore')
                #ws.cell(row=row_no, column=2, value=web_link)
                cell = WriteOnlyCell(ws, value=link.text)
                ws.append([cell, web_link ])
                row_no +=1
                print(link.text + ":" + link.get_attribute('href'))
    print("Total Records:", row_no)
    wb.save("All UK HE Providers.xlsx")
    driver.quit()

def getMainPage():
    page_link = []
    print("main page")
    driver = webdriver.Chrome("C:/chromedriver_win32/chromedriver.exe")
    driver.get("http://www.hefce.ac.uk/reg/register/search/Browse")
    print(driver.title)
    for li in driver.find_elements_by_xpath('//div[@class="atozlist"]//ul[@id="glossaryaz"]//li'):
        anchors = li.find_elements_by_tag_name('a')
        for a in anchors:
            link= driver.find_element_by_link_text(a.text)
            page_link.append(unicodedata.normalize('NFKD', link.get_attribute('href')).encode('ascii','ignore'))

    driver.quit()
    time.sleep(10)
    getDataPerpage(page_link)

if __name__=="__main__":
    getMainPage()